from dateutil import parser
import datetime as dt
import urllib.error

from openpyxl import load_workbook
from TogglPy import Toggl
from TogglPy import Endpoints

import config


class GermanParserInfo(parser.parserinfo):
    WEEKDAYS = [("Mo", "Montag"),
                ("Di", "Dienstag"),
                ("Mi", "Mittwoch"),
                ("Do", "Donnerstag"),
                ("Fr", "Freitag"),
                ("Sa", "Samstag"),
                ("So", "Sonntag")]
    MONTHS = [("Jän", "Jänner"),
              ("Feb", "Februar"),
              ("März", "März"),
              ("Apr", "April"),
              ("Mai", "May"),
              ("Juni", "June"),
              ("Juli", "July"),
              ("Aug", "August"),
              ("Sept", "September"),
              ("Okt", "Oktober"),
              ("Nov", "November"),
              ("Dez", "Dezember")]

    def weekday(self, name):
        try:
            return self._weekdays[name.lower()]
        except KeyError:
            pass
        return None

    def day(self, weekday):
        for name, id in self._weekdays.items():
            if len(name) > 2 and weekday == id:
                return name

    def shortday(self, weekday):
        for name, id in self._weekdays.items():
            if len(name) == 2 and weekday == id:
                return name

dateparser = GermanParserInfo()


def weekdays_consecutive(days):
    indexes = [dateparser.weekday(d) for d in days]
    indexes.sort()
    expected = range(indexes[0], indexes[-1] + 1)
    match = [i for i, j in zip(indexes, expected) if i == j]
    return len(match) == len(indexes)

# Get the information about missing dates
wb2 = load_workbook(config.TIMESHEET)
names = wb2.get_sheet_names()
init = config.INIT


def hour_range_to_seconds(info):
    hours = info.split('-')
    start = parser.parse(hours[0])
    end = parser.parse(hours[1])
    duration = end - start
    return duration.total_seconds()


def calculate_soll_seit(info):
    seconds = hour_range_to_seconds(info)
    hh = int(seconds // 3600)
    mm = int((seconds % 3600) // 60)
    return '{:02d}:{:02d}'.format(hh, mm)

# This might be better if as a method, triggered by a launch argument, the value
# shuld not go in the configuration file
if init:
    # Befüllung der Normdienstzeiten
    # Find days with common hours, group them. Leave the others
    common = dict()
    soll_seit = dict()
    for day, info in config.WORKING_SCHEDULE_PER_DAY.items():
        soll_seit[day] = calculate_soll_seit(info)
        try:
            common[info].append(day)
        except KeyError:
            common[info] = [day]
    # Group the values for each time slot by range, e.g. Mo-Do
    for slot, days in common.items():
        if len(days) > 2:
            if weekdays_consecutive(days):
                days.sort(key=dateparser.weekday)
                common[slot] = [days[0], days[-1]]
    # Fill Regeldienstzeit
    # FIXME wat is Blockzeiten
    for name in names[1:]:  # Skip the guide
        ws = wb2.get_sheet_by_name(name)
        ranges = [": ".join(["-".join(val), key]) for key, val in common.items()]
        ws['A3'] = 'Regeldienstzeit: {}'.format(
            ", ".join(ranges))
        # Fill Soll Zeit
        days = ws['A'][5:]
        for day in days:
            if day.value is None:
                break
            # Add info if not bank holiday (Color is red)
            value = None
            if day.value in config.WORKING_SCHEDULE_PER_DAY:
                if day.font.color is None:      # Black is default, so there is no color info for regular days
                    value = soll_seit[day.value]
            ws['H{}'.format(day.row)] = value   # Make sure any wrong inputs are cleared

    wb2.save(config.TIMESHEET)


def get_start_date(datum_column, ws):
    for datum in datum_column:
        if datum.value is None:
            return None, None
        if datum.font.color is None:  # Black is default, so there is no color info for regular days
            von1 = ws['C{}'.format(datum.row)]
            if von1.value is None:
                return datum.value, datum.row

start_date = None
sheet = None
for name in names[1:]:      # Skip the guide
    sheet = wb2.get_sheet_by_name(name)
    colB = sheet['B'][5:]
    start_date, row = get_start_date(colB, sheet)
    if start_date is not None:
        break

expected_duration = dict()

for day, info in config.WORKING_SCHEDULE_PER_DAY.items():
    expected_duration[day] = hour_range_to_seconds(info)

if start_date is None:
    print('No missing data in the sheet.')
else:
    toggl = Toggl()
    toggl.setAPIKey(config.API_TOKEN)
    data = dict()
    data['user_agent'] = config.EMAIL
    # For each missing date, 1. Get start hour, 2. Get total hours (this might span several records in ALL workspaces
    start_duration = dict()
    start_clockin = dict()
    for ws, info in config.WORK_SPACES.items():
        # Get the client ids
        client_ids = []
        clients = toggl.getClients()  # get all clients
        for client in clients:  # search through them for one matching the name provided
            if client['name'] in info['clients']:
                client_ids.append(str(client['id']))
        data['client_ids'] = ','.join(client_ids)
        data['project_ids'] = ','.join(info['projects'])
        data['workspace_id'] = ws
        data['since'] = start_date
        data['until'] = dt.datetime.today() - dt.timedelta(days=1)
        try:
            response = toggl.getDetailedReport(data)
            print(response)
            # We assume no pages (i.e. update is done regularly
            entries = response['data']
            for entry in entries:
                start = parser.parse(entry['start'])
                start = str(start.date())
                duration = int(entry['dur'])
                try:
                    start_duration[start] += duration
                    if start_clockin[start] > parser.parse(entry['start']).time():
                        start_clockin[start] = parser.parse(entry['start']).time()
                except KeyError:
                    start_duration[start] = duration
                    start_clockin[start] = parser.parse(entry['start']).time()
        except urllib.error.HTTPError as e:
            print(e.msg)

    # For each date, 1. Find the expected rage with weekday
    # If range is covered set an end time to match range
    # if range is not covered set an end time to match duration (warn the user about under work)
    for start, duration in start_duration.items():
        date = parser.parse(start)
        row = 5 + date.day
        weekday = date.weekday()
        shortday = dateparser.shortday(weekday).capitalize()
        try:
            hour_range = config.WORKING_SCHEDULE_PER_DAY[shortday]
        except KeyError:  # Not a working day
            continue
        day_duration = hour_range_to_seconds(hour_range)
        von1 = start_clockin[start]
        dt_von1 = dt.datetime.combine(dt.date(1, 1, 1), von1)
        if duration > (day_duration * 1000):
            # bis1 = von1 + time(second=int(day_duration/1000))
            bis1 = (dt_von1 + dt.timedelta(seconds=day_duration)).time()
        else:
            # bis1 = von1 + time(second=int(duration/1000))
            bis1 = (dt_von1 + dt.timedelta(milliseconds=duration)).time()
        print(von1, bis1)
        sheet['C{}'.format(row)] = von1.replace(second=0, microsecond=0)
        sheet['D{}'.format(row)] = bis1.replace(second=0, microsecond=0)
    wb2.save(config.TIMESHEET)